import re
import sys

import pandas as pd
import datetime as datetime
from pytz import timezone

import smtplib
from openpyxl import load_workbook

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

pd.options.display.max_columns = None
pd.options.display.max_rows = None

import datetime
import base64

from msal import ConfidentialClientApplication

sys.path.insert(0, 'C:/Users/Administrator/AdQvestDir/Adqvest_Function')
import adqvest_db
import JobLogNew as log
from AdqvestEmailSender import adqvestemailsender

from clickhouse_driver import Client
host = '172.31.67.17'
user_name = 'default'
password_string = 'Clickhouse@2025'
db_name = 'AdqvestDB'
client = Client(host, user=user_name, password=password_string, database=db_name)

india_time = timezone('Asia/Kolkata')
today      = datetime.datetime.now(india_time)
days       = datetime.timedelta(1)
yesterday = today - days

run_time = pd.to_datetime(today.strftime("%Y-%m-%d %H:%M:%S"))

def run_program(run_by='Adqvest_Bot', py_file_name=None):
    engine = adqvest_db.db_conn()
    india_time = timezone('Asia/Kolkata')

    ## job log details
    job_start_time = datetime.datetime.now(india_time)
    table_name = ''
    if (py_file_name is None):
        py_file_name = sys.argv[0].split('.')[0]
    scheduler = ''
    no_of_ping = 0
    
    try :
        if(run_by=='Adqvest_Bot'):
            log.job_start_log_by_bot(table_name,py_file_name,job_start_time)
        else:
            log.job_start_log(table_name,py_file_name,job_start_time,scheduler)

        # Define multiple recipients
        to = 'mrinalini@thurro.com,karthik@thurro.com'
        cc = 's.heteshkumar@thurro.com,gokulakrishnan.j@thurro.com'
        print('CC Updated')

        # msg = MIMEMultipart()
        # msg['From'] = "adqvest.insights@thurro.com"
        # msg['To'] = ", ".join(to)  # Join multiple recipients with comma
        # msg['Cc'] = ", ".join(cc)  # Join CC recipients with comma

        # # Combine all recipients for smtp.sendmail()
        # recipients = to + cc

        # msg['Subject'] = "Company Count of Inv. Presentation and Conf. Call documents as on " + str(today.date())
        # body = "Hello , Please find the Excel Sheet Attached with this Email" 
        # msg.attach(MIMEText(body, 'plain'))

        sel_time = (datetime.datetime.now(india_time) - datetime.timedelta(1)).strftime("%Y-%m-%d %H:%M:%S")
        sel_end_time = (datetime.datetime.now(india_time) - datetime.timedelta(hours = 2)).strftime("%Y-%m-%d %H:%M:%S")
        time_now = pd.to_datetime(today.strftime('%Y-%m-%d %H:%M:%S'))

        # NSE data

        Ch = '''WITH
                -- Step 1: All distinct companies in lowercase
                all_companies AS (
                    SELECT DISTINCT lower(Document_Company) AS company_lower
                    FROM NSE_INVESTOR_INFORMATION_DAILY_DATA_CORPUS_CHUNKED
                ),
                
                -- Step 2: Cross join companies with the 3 subject types
                company_subjects AS (
                    SELECT
                        company_lower,
                        subject
                    FROM all_companies
                    CROSS JOIN (
                        SELECT 'Investor Presentation' AS subject
                        UNION ALL
                        SELECT 'Conference Call'
                        UNION ALL
                        SELECT 'Press Release'
                    ) s
                ),
                
                -- Step 3: All docs with subject tag
                all_docs AS (
                    SELECT
                        lower(Document_Company) AS company_lower,
                        Document_Content,
                        Document_Year,
                        Document_Id,
                        CASE
                            WHEN lower(Document_Content) ILIKE '%investor presentation%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%results presentation%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%results performance%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%result presentation%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%result performance%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%earnings presentation%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%earning performance%' THEN 'Investor Presentation'
                            WHEN lower(Document_Content) ILIKE '%press releases%' THEN 'Press Release'
                            WHEN lower(Document_Content) ILIKE '%press release%' THEN 'Press Release'
                            WHEN lower(Document_Content) ILIKE '%con. call%' 
                                 OR lower(Document_Content) LIKE '%con. calls%' 
                                 OR lower(Document_Content) LIKE '%conference call%' THEN 'Conference Call'
                            ELSE NULL
                        END AS subject
                    FROM NSE_INVESTOR_INFORMATION_DAILY_DATA_CORPUS_CHUNKED
                ),
                
                -- Step 4: Final join
                final_docs AS (
                    SELECT
                        cs.company_lower AS Document_Company,
                        cs.subject,
                        d.Document_Year,
                        d.Document_Id
                    FROM company_subjects cs
                    LEFT JOIN all_docs d
                        ON cs.company_lower = d.company_lower AND cs.subject = d.subject
                )

            -- Step 5: Pivot counts
            SELECT
                Document_Company AS company,
                subject,
                NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q1 FY25'), 0) AS `Q1 FY25`,
                NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q2 FY25'), 0) AS `Q2 FY25`,
                NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q3 FY25'), 0) AS `Q3 FY25`,
                NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q4 FY25'), 0) AS `Q4 FY25`
            FROM final_docs
            GROUP BY Document_Company, subject
            ORDER BY Document_Company, subject;

        '''

        # Ch = '''WITH
        #         -- Step 1: All distinct companies in lowercase
        #         all_companies AS (
        #             SELECT DISTINCT lower(Document_Company) AS company_lower
        #             FROM NSE_INVESTOR_INFORMATION_PAST_DATA_CORPUS_CHUNKED
        #             UNION ALL
        #             SELECT DISTINCT lower(Document_Company) AS company_lower
        #             FROM NSE_INVESTOR_INFORMATION_DAILY_DATA_CORPUS_CHUNKED
        #         ),
                
                # -- Step 2: Cross join companies with the 2 subject types
                # company_subjects AS (
                #     SELECT
                #         company_lower,
                #         subject
                #     FROM all_companies
                #     CROSS JOIN (
                #         SELECT 'Investor Presentation' AS subject
                #         UNION ALL
                #         SELECT 'Conference Call'
                #         UNION ALL
                #         SELECT 'Press Release'
                #     ) s
                # ),
                
                # -- Step 3: All docs with subject tag (searching in Document_Content)
                # all_docs AS (
                #     SELECT
                #         lower(Document_Company) AS company_lower,
                #         Document_Content,
                #         Document_Year,
                #         Document_Id,
                #         CASE
                #             WHEN lower(Document_Content) ILIKE '%investor presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%results presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%results performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%result presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%result performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%earnings presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%earning performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%press releases%' THEN 'Press Release'
                #             WHEN lower(Document_Content) ILIKE '%press release%' THEN 'Press Release'
                #             WHEN lower(Document_Content) ILIKE '%con. call%' OR lower(Document_Content) LIKE '%con. calls%' OR lower(Document_Content) LIKE '%conference call%' THEN 'Conference Call'
                #             ELSE NULL
                #         END AS subject
                #     FROM NSE_INVESTOR_INFORMATION_PAST_DATA_CORPUS_CHUNKED
                #     UNION ALL
                #     SELECT
                #         lower(Document_Company) AS company_lower,
                #         Document_Content,
                #         Document_Year,
                #         Document_Id,
                #         CASE
                #             WHEN lower(Document_Content) ILIKE '%investor presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%results presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%results performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%result presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%result performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%earnings presentation%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%earning performance%' THEN 'Investor Presentation'
                #             WHEN lower(Document_Content) ILIKE '%press releases%' THEN 'Press Release'
                #             WHEN lower(Document_Content) ILIKE '%press release%' THEN 'Press Release'
                #             WHEN lower(Document_Content) ILIKE '%con. call%' OR lower(Document_Content) LIKE '%con. calls%' OR lower(Document_Content) LIKE '%conference call%' THEN 'Conference Call'
                #             ELSE NULL
                #         END AS subject
                #     FROM NSE_INVESTOR_INFORMATION_DAILY_DATA_CORPUS_CHUNKED
                # ),
                
                # -- Step 4: Final join to link all companies + subjects with docs (LEFT JOIN keeps all companies/subjects)
                # final_docs AS (
                #     SELECT
                #         cs.company_lower AS Document_Company,
                #         cs.subject,
                #         d.Document_Year,
                #         d.Document_Id
                #     FROM company_subjects cs
                #     LEFT JOIN all_docs d
                #         ON cs.company_lower = d.company_lower AND cs.subject = d.subject
                # )
                
                # -- Step 5: Pivot counts for Q1 FY25 to Q4 FY25
                # SELECT
                #     Document_Company AS company,
                #     subject,
                #     NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q1 FY25'), 0) AS `Q1 FY25`,
                #     NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q2 FY25'), 0) AS `Q2 FY25`,
                #     NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q3 FY25'), 0) AS `Q3 FY25`,
                #     NULLIF(countDistinctIf(Document_Id, Document_Year = 'Q4 FY25'), 0) AS `Q4 FY25`
                # FROM final_docs
                # GROUP BY Document_Company, subject
                # ORDER BY Document_Company, subject;
                # ''' 

        a,cols = client.execute(Ch,with_column_types=True)
        CH_df = pd.DataFrame(a, columns=[tuple[0] for tuple in cols])
        print('DF Generated')

        CH_df['company'] = CH_df['company'].str.title()
        CH_df.rename(columns={'company': 'Company','subject':'Subject'}, inplace=True)

        df_melted = CH_df.melt(id_vars=['Company', 'Subject'], var_name='Quarter', value_name='Value')

        pivot_df = df_melted.pivot_table(
                index='Company',
                columns=['Quarter', 'Subject'],
                values='Value',
                aggfunc='sum'
            )
        
        pivot_df = pivot_df.sort_index(axis=1, level=0)
        
        import numpy as np
        pivot_df = pivot_df.replace(0, np.nan)
        
        excel_path = 'NSE_COMPANYWISE_INV_PRESENTATION_CONF_CALL_FY25.xlsx'
        pivot_df.to_excel(excel_path)
        
        company_list = df_melted['Company'].str.strip().str.lower().unique().tolist()
                
        events = pd.read_sql("SELECT * FROM NSE_INVESTOR_INFORMATION_EVENTS_CALENDAR_RANDOM_DATA", engine)

        events.columns = events.columns.str.strip()

        filtered_df = events[events['Details'].str.contains('March 31, 2025', case=False, na=False)]

        # Step 1: Standardize company names to lowercase for matching
        df_melted['Company'] = df_melted['Company'].str.strip().str.lower()
        filtered_df['Company_Name'] = filtered_df['Company_Name'].str.strip().str.lower()

        # Step 2: Filter df_melted to only Q4 FY25
        q4_data = df_melted[df_melted['Quarter'] == 'Q4 FY25']

        # Step 3: Get unique companies that appear in the event calendar (filings)
        filed_companies = filtered_df['Company_Name'].unique()

        # Step 4: For those companies, check if any subject (Conf Call / Inv Pres) has NaN or 0
        q4_subset = q4_data[q4_data['Company'].isin(filed_companies)]

        # Step 5: Identify companies where ALL subjects have 0 or NaN (i.e., missed everything)
        summary = (
            q4_subset
            .groupby('Company')['Value']
            .apply(lambda x: all(pd.isna(x) | (x == 0)))
            .reset_index(name='missed_all')
        )

        # Step 6: Filter only the missed ones
        missed_companies_df = summary[summary['missed_all']]

        wb = load_workbook(excel_path)
        ws = wb.active

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)  # Safe even for merged cells
        
            for cell in column_cells:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
        
            ws.column_dimensions[col_letter].width = max_length + 2  # add padding
            
        from openpyxl.styles import PatternFill
        
        red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")

        missing_companies_set = set(missed_companies_df['Company'].str.strip().str.lower())
        
        for row in ws.iter_rows(min_row=2):
            company_cell = row[0]  # Assuming company name is in the first column
            if company_cell.value and str(company_cell.value).strip().lower() in missing_companies_set:
                for cell in row:
                    cell.fill = red_fill    

        wb.save(excel_path)
            
        # Attach Count file
        # part = MIMEBase('application', "octet-stream")
        # part.set_payload(open("NSE_COMPANYWISE_INV_PRESENTATION_CONF_CALL_FY25.xlsx", "rb").read())
        # encoders.encode_base64(part)
        # part.add_header('Content-Disposition', 'attachment; filename="NSE_COMPANYWISE_INV_PRESENTATION_CONF_CALL_FY25.xlsx"')
        # msg.attach(part)

        # Send email
        email_sender = adqvestemailsender()

        body_parts=[]

        body ="""Hello , Please find the Excel Sheet Attached with this Email."""
        body_parts.append({'type': 'plain', 'content': body})

        subject = "Company Count of Inv. Presentation and Conf. Call documents as on " + str(today.date())

        message = email_sender.create_email_message(to, cc, subject, body_parts)

        with open("NSE_COMPANYWISE_INV_PRESENTATION_CONF_CALL_FY25.xlsx", "rb") as f:
            excel_bytes = f.read()

        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')

        attachment = {
            "@odata.type": "#microsoft.graph.fileAttachment","name": "NSE_COMPANYWISE_INV_PRESENTATION_CONF_CALL_FY25.xlsx",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","contentBytes": excel_base64}

        if "attachments" not in message["message"]:
            message["message"]["attachments"] = []
        message["message"]["attachments"].append(attachment)
        
        if email_sender.send_email(message):
            print("Email sent successfully via Microsoft Graph API")
        else:
            print("Failed to send email via Microsoft Graph API")

        # smtp = smtplib.SMTP('smtp.gmail.com:587')
        # smtp.starttls()
        # smtp.login("thurro@adqvest.com", "yzjw egjr ryuc namr")
        # smtp.sendmail(msg['From'], recipients, msg.as_string())
        # smtp.quit()
        print('REPORT SENT')

        log.job_end_log(table_name,job_start_time, no_of_ping)
    except Exception as e:
        error_type = str(re.search("'(.+?)'",str(sys.exc_info()[0])).group(1))
        error_msg = str(e) + " line " + str(sys.exc_info()[2].tb_lineno)
        print(error_msg)
        log.job_error_log(table_name,job_start_time,error_type,error_msg, no_of_ping)

if(__name__=='__main__'):
    run_program(run_by='manual')
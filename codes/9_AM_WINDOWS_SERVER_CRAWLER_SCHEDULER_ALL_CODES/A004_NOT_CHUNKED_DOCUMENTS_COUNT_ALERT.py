# -*- coding: utf-8 -*-
"""
Created on Tue Jul 29 08:57:10 2025

@author: GOKUL
"""

import re
import sys

import pandas as pd
import datetime as datetime
from pytz import timezone

from openpyxl import load_workbook
import base64

sys.path.insert(0, 'C:/Users/Administrator/AdQvestDir/Adqvest_Function')
import adqvest_db
import JobLogNew as log
from AdqvestEmailSender import adqvestemailsender

india_time = timezone('Asia/Kolkata')
today      = datetime.datetime.now(india_time)
days       = datetime.timedelta(1)
yesterday = today - days

run_time = pd.to_datetime(today.strftime("%Y-%m-%d %H:%M:%S"))

def run_program(run_by='Adqvest_Bot', py_file_name=None):
    engine = adqvest_db.db_conn()
    india_time = timezone('Asia/Kolkata')

    ## job log details
    job_start_time = datetime.datetime.now(india_time)
    table_name = ''
    if (py_file_name is None):
        py_file_name = sys.argv[0].split('.')[0]
    scheduler = ''
    no_of_ping = 0
    
    try :
        if(run_by=='Adqvest_Bot'):
            log.job_start_log_by_bot(table_name,py_file_name,job_start_time)
        else:
            log.job_start_log(table_name,py_file_name,job_start_time,scheduler)

        # Define multiple recipients
        to = 'mrinalini@thurro.com,karthik@thurro.com,induja.p@thurro.com,  s.heteshkumar@thurro.com, santonu.debnath@thurro.com,gokulakrishnan.j@thurro.com'
        cc = 's.akhuli@thurro.com'
        print('Recipients Updated')
               
        tables = ['CARE_RATINGS_DAILY_DATA_CORPUS','CRISIL_DAILY_DATA_CORPUS','ICRA_DAILY_DATA_CORPUS',
                  'INDIA_BUDGET_SPEECHES_YEARLY_DATA_CORPUS', 'INDIA_UNION_BUDGET_YEARLY_DATA_CORPUS',
                  'INDIA_BUDGET_ECONOMIC_SURVEY_YEARLY_DATA_CORPUS',
                  'NSE_MARKET_PULSE_MONTHLY_DATA_CORPUS',
                  'SEBI_FINAL_OFFER_DOCUMENTS_RHP_DRHP_DAILY_DATA_CORPUS',
                  'PIB_REPORTS_DAILY_DATA_CORPUS',
                  'ANAROCK_MARKET_VIEWPOINTS_MONTHLY_DATA_CORPUS',
                  'BSE_ANNOUNCEMENT_ANNUAL_REPORT_COMPANY_WISE_YEARLY_DATA_CORPUS',
                  'CAG_STATE_WISE_ACCOUNT_MONTHLY_DATA_CORPUS','CAG_AUDIT_REPORTS_RANDOM_DATA_CORPUS',
                  'RBI_PRESS_RELEASES_DAILY_DATA_CORPUS','RBI_ANNUAL_REPORT_YEARLY_DATA_CORPUS',
                  'NSE_INVESTOR_INFORMATION_DAILY_DATA_CORPUS', 'NSE_DOC_LINKS_FROM_CORPUS_DAILY_DATA',
                  'DEA_ECONOMY_REPORT_MONTHLY_DATA_CORPUS']
        
        chunked_table_lengths = []
        
        for sql_Table in tables:
            df_links=pd.read_sql(f''' 
                                 SELECT
                                    count(*) AS Total_Count,
                                    
                                    SUM(
                                        CASE 
                                            WHEN S3_Upload_Status = 'Done' 
                                                and (Chunking_Status is Null or Chunking_Status != 'Not Relevant') 
                                            THEN 1 ELSE 0
                                        END
                                    ) AS Total_Relevant_Count,
                                    
                                    -- Not Chunked
                                    SUM(
                                        CASE
                                            WHEN Chunking_Status IS NULL 
                                                 AND Chunking_Comments IS NULL 
                                                 AND S3_Upload_Status = 'Done'
                                            THEN 1 ELSE 0
                                        END
                                    ) AS Not_Chunked_Count,
                                    
                                    -- Chunked
                                    SUM(
                                        CASE
                                            WHEN Chunking_Status = 'Done'
                                                 AND S3_Upload_Status = 'Done'
                                            THEN 1 ELSE 0
                                        END
                                    ) AS Chunked_Count,
                                                                        
                                    MIN(DATE(Runtime)) AS Min_Runtime,
                                    MAX(DATE(Runtime)) AS Max_Runtime,
                                    MAX(Runtime) AS Last_Runtime
                            
                                FROM AdqvestDB.{sql_Table}
                                where DATE(Runtime) < CURRENT_DATE
                            ''', engine)
            
            # chunked_table_lengths.append({'Table_Name': sql_Table,'Row_Count': int(df_links['NotDownloaded_Count'].iloc[0]), 'Min_Date': df_links['Min_Runtime'].iloc[0],'Max_Date': df_links['Max_Runtime'].iloc[0] })
            
            chunked_table_lengths.append({'Table_Name': sql_Table, 
                                          'Total_Count': int(df_links['Total_Count'].iloc[0]),
                                          'Total_Relevant_Count': int(df_links['Total_Relevant_Count'].iloc[0]),
                                          'Chunked_Count': int(df_links['Chunked_Count'].iloc[0]),
                                          'Not_Chunked_Count': int(df_links['Not_Chunked_Count'].iloc[0]), 
                                          'Min_Date': df_links['Min_Runtime'].iloc[0],
                                          'Max_Date': df_links['Max_Runtime'].iloc[0], 
                                          'Last_Runtime': df_links['Last_Runtime'].iloc[0]})

        chunked_lengths_df = pd.DataFrame(chunked_table_lengths)
        # chunked_lengths_df = chunked_lengths_df[chunked_lengths_df['Not_Chunked_Count'] != 0]
        chunked_lengths_df.reset_index(drop=True,inplace=True)
                
        print('DF Generated')

        excel_path = 'NOT_CHUNKED_TABLES_STATUS.xlsx'
        chunked_lengths_df.to_excel(excel_path, index=False)

        wb = load_workbook(excel_path)
        ws = wb.active

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)  # Safe even for merged cells
        
            for cell in column_cells:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
        
            ws.column_dimensions[col_letter].width = max_length + 2  # add padding
            
        wb.save(excel_path)
            
        # Send email
        email_sender = adqvestemailsender()

        body_parts=[]

        body ="""Hello , Please find the Excel Sheet Attached with this Email."""
        body_parts.append({'type': 'plain', 'content': body})

        subject = "Count of documents that were not chunked as of " + str(today.date())

        message = email_sender.create_email_message(to, cc, subject, body_parts)

        with open("NOT_CHUNKED_TABLES_STATUS.xlsx", "rb") as f:
            excel_bytes = f.read()

        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')

        attachment = {
            "@odata.type": "#microsoft.graph.fileAttachment","name": "NOT_CHUNKED_TABLES_STATUS.xlsx",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","contentBytes": excel_base64}

        if "attachments" not in message["message"]:
            message["message"]["attachments"] = []
        message["message"]["attachments"].append(attachment)
        
        if email_sender.send_email(message):
            print("Email sent successfully via Microsoft Graph API")
        else:
            print("Failed to send email via Microsoft Graph API")

        print('REPORT SENT')

        log.job_end_log(table_name,job_start_time, no_of_ping)
    except Exception as e:
        error_type = str(re.search("'(.+?)'",str(sys.exc_info()[0])).group(1))
        error_msg = str(e) + " line " + str(sys.exc_info()[2].tb_lineno)
        print(error_msg)
        log.job_error_log(table_name,job_start_time,error_type,error_msg, no_of_ping)

if(__name__=='__main__'):
    run_program(run_by='manual')